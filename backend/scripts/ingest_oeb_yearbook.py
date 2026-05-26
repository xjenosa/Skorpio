"""
Ingest the Ontario Energy Board's Yearbook of Electricity Distributors
(open-data XLSX files) and emit per-utility customer counts, winter peak
load, and reliability indices for every Ontario LDC.

Sources (latest published year: 2021):
  - yearbook-General-Statistics-2021.xlsx
      Sheet 1 = customer counts per rate class
      Sheet 2 = peak load (kW), service area, line kilometres
  - yearbook-System-Reliability-2021.xlsx
      Sheet 1 = SAIFI / SAIDI per cause code

Used to overlay real numbers onto `feeder_topology.CITY_PROFILES` and
`utility_assets._REGISTRY` for Ontario utilities (Toronto Hydro, Alectra,
Hydro Ottawa, Hydro One Networks, etc.).

Usage:

    python -m backend.scripts.ingest_oeb_yearbook

Caches:
  data/cache/oeb-general-stats-2021.xlsx
  data/cache/oeb-reliability-2021.xlsx
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path
from urllib.request import urlopen

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
CACHE_DIR = REPO_ROOT / "data" / "cache"
OUTPUT_FILE = REPO_ROOT / "backend" / "services" / "_oeb_utilities_generated.py"

OEB_BASE = "https://www.oeb.ca/sites/default/files"
GENERAL_URL = f"{OEB_BASE}/yearbook-General-Statistics-2021.xlsx"
RELIABILITY_URL = f"{OEB_BASE}/yearbook-System-Reliability-2021.xlsx"
GENERAL_CACHE = CACHE_DIR / "oeb-general-stats-2021.xlsx"
RELIABILITY_CACHE = CACHE_DIR / "oeb-reliability-2021.xlsx"

TARGET_YEAR = 2021  # latest year the OEB has published as of this writing

# Sheet 1's "Customer_or_Connections" column separates real metered customers
# (Residential, General Service, Large User, Sub Transmission) from raw
# connection counts (street lights, embedded distributors, unmetered scatter).
# The customer-count we want is just the "Customers" half — that's what every
# OEB scorecard treats as "total customers".
CUSTOMERS_FLAG = "Customers"


# ── Download ────────────────────────────────────────────────────────────── #

def _download_if_missing(url: str, dest: Path) -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 50_000:
        print(f"[ok] Using cached {dest.name} ({dest.stat().st_size / 1e3:.0f} KB)")
        return dest
    print(f"[..] Downloading {url} ...")
    with urlopen(url) as resp, dest.open("wb") as out:
        while chunk := resp.read(1 << 16):
            out.write(chunk)
    print(f"[ok] Saved to {dest} ({dest.stat().st_size / 1e3:.0f} KB)")
    return dest


# ── Parse general statistics ────────────────────────────────────────────── #

def _f(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_general(path: Path) -> dict[str, dict]:
    """Returns {company_name: {customers, winter_peak_mw, summer_peak_mw,
    service_area_km2, line_km}} for TARGET_YEAR."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: dict[str, dict] = defaultdict(dict)

    # Sheet 1 — sum customer counts across useful rate classes per company.
    customers: dict[str, float] = defaultdict(float)
    ws = wb["Sheet 1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    for row in rows:
        try:
            year = int(_f(row[idx["Year"]]) or 0)
        except (TypeError, ValueError):
            continue
        if year != TARGET_YEAR:
            continue
        if row[idx["Customer_or_Connections"]] != CUSTOMERS_FLAG:
            continue
        company = (row[idx["Company_Name"]] or "").strip()
        count = _f(row[idx["Total_Customers_or_Connections"]]) or 0
        customers[company] += count

    # Sheet 2 — pick the row for TARGET_YEAR per company; convert kW to MW.
    ws = wb["Sheet 2"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    for row in rows:
        try:
            year = int(_f(row[idx["Year"]]) or 0)
        except (TypeError, ValueError):
            continue
        if year != TARGET_YEAR:
            continue
        company = (row[idx["Company_Name"]] or "").strip()
        winter_kw = _f(row[idx["Winter_Peak_Load_With_Embedded_Generation_kW"]])
        summer_kw = _f(row[idx["Summer_Peak_Load_With_Embedded_Generation_kW"]])
        area = _f(row[idx["Service_Area_Total_Square_Kilometers"]])
        line = _f(row[idx["Total_Circuit_Kilometers_of_Line"]])
        out[company] = {
            "winter_peak_mw": round(winter_kw / 1000.0, 1) if winter_kw else None,
            "summer_peak_mw": round(summer_kw / 1000.0, 1) if summer_kw else None,
            "service_area_km2": round(area, 1) if area else None,
            "line_km": round(line, 1) if line else None,
        }

    # Merge customers in. Sheet 2 entries may not have a Sheet 1 match (rare).
    for company, n in customers.items():
        out.setdefault(company, {})["customers"] = int(n)
    for company in out:
        out[company].setdefault("customers", None)
    return dict(out)


# ── Parse reliability ───────────────────────────────────────────────────── #

def parse_reliability(path: Path) -> dict[str, dict]:
    """Returns {company_name: {saifi, saidi}} for TARGET_YEAR, aggregated
    across all cause codes."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet 1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}

    agg: dict[str, dict] = defaultdict(lambda: {"saifi": 0.0, "saidi": 0.0,
                                                  "customers_served": 0.0,
                                                  "cause_rows": 0})
    for row in rows:
        try:
            year = int(_f(row[idx["Year"]]) or 0)
        except (TypeError, ValueError):
            continue
        if year != TARGET_YEAR:
            continue
        company = (row[idx["Company_Name"]] or "").strip()
        bucket = agg[company]
        # SAIFI / SAIDI in the XLSX are reported per cause code; summing them
        # across causes gives the utility's total annual figure (this is how
        # the OEB itself rolls them up in the scorecard).
        saifi = _f(row[idx["SAIFI_from_Major_Events"]]) or 0.0
        saidi = _f(row[idx["SAIDI_from_Major_Events"]]) or 0.0
        bucket["saifi"] += saifi
        bucket["saidi"] += saidi
        served = _f(row[idx["Average_Number_of_Customers_Served"]])
        if served and served > bucket["customers_served"]:
            bucket["customers_served"] = served
        bucket["cause_rows"] += 1

    # Drop the helper counter; round results.
    out: dict[str, dict] = {}
    for company, b in agg.items():
        out[company] = {
            "saifi": round(b["saifi"], 3),
            "saidi": round(b["saidi"], 3),
        }
    return out


# ── Emit ────────────────────────────────────────────────────────────────── #

HEADER = '''"""
Auto-generated per-utility data for Ontario electricity distributors from the
OEB Yearbook open-data files (2021, latest published).

Sources:
  - https://www.oeb.ca/sites/default/files/yearbook-General-Statistics-2021.xlsx
  - https://www.oeb.ca/sites/default/files/yearbook-System-Reliability-2021.xlsx

Each entry merges:
  - General Statistics: customer count (sum of useful rate classes),
    winter/summer peak load (MW), total service area (km²), total circuit km
  - System Reliability: SAIFI + SAIDI summed across cause codes

DO NOT EDIT BY HAND - re-run `python -m backend.scripts.ingest_oeb_yearbook`.

Used by:
  - feeder_topology.py CITY_PROFILES (real customer counts + winter peak for
    Toronto Hydro, Alectra, Hydro Ottawa)
  - utility_assets.py (sanity-check on synthesised asset counts)
"""

OEB_UTILITIES_BY_NAME: dict[str, dict] = {
'''


def emit_python(merged: dict[str, dict], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        f.write(HEADER)
        for company in sorted(merged):
            e = merged[company]
            f.write(f'    {company!r}: {{\n')
            for key in ("customers", "winter_peak_mw", "summer_peak_mw",
                        "service_area_km2", "line_km", "saifi", "saidi"):
                v = e.get(key)
                if v is None:
                    f.write(f'        "{key}": None,\n')
                elif isinstance(v, int):
                    f.write(f'        "{key}": {v},\n')
                else:
                    f.write(f'        "{key}": {float(v):.3f},\n')
            f.write(f'    }},\n')
        f.write("}\n")
    print(f"  Wrote {path}")


def main() -> None:
    gen_path = _download_if_missing(GENERAL_URL, GENERAL_CACHE)
    rel_path = _download_if_missing(RELIABILITY_URL, RELIABILITY_CACHE)
    gen = parse_general(gen_path)
    rel = parse_reliability(rel_path)

    merged: dict[str, dict] = {}
    for company in set(gen) | set(rel):
        e = dict(gen.get(company, {}))
        e.update(rel.get(company, {}))
        merged[company] = e
    print(f"  Merged {len(merged)} utilities.")
    emit_python(merged, OUTPUT_FILE)

    print(f"\nSpot-check (largest 5 by customer count, {TARGET_YEAR}):")
    ranked = sorted(merged.items(), key=lambda kv: -(kv[1].get("customers") or 0))[:5]
    for company, e in ranked:
        print(f"  {company[:50]:50s} cust={e.get('customers'):>10},  "
              f"winter={e.get('winter_peak_mw')} MW,  "
              f"SAIFI={e.get('saifi')},  SAIDI={e.get('saidi')}")


if __name__ == "__main__":
    main()
